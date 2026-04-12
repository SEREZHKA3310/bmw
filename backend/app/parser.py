import io
from openpyxl import load_workbook


AGREED_SHEET_NAME = "Согласованные оценки"
PROFILE_SHEET_NAME = "Профиль"


def as_float(value):
    try:
        if value is None or value == "":
            return 0.0
        return round(float(str(value).replace(",", ".")), 2)
    except:
        return 0.0


def clean_text(value):
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split()).strip()


def parse_agreed_sheet(wb):
    if AGREED_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f'Лист "{AGREED_SHEET_NAME}" не найден')

    ws = wb[AGREED_SHEET_NAME]

    competences = []
    indicators = []

    for row_idx in range(4, ws.max_row + 1):
        title = clean_text(ws.cell(row=row_idx, column=1).value)

        if not title:
            continue

        if title.startswith("Далее:"):
            break

        employee_score = as_float(ws.cell(row=row_idx, column=13).value)
        manager_score = as_float(ws.cell(row=row_idx, column=14).value)
        agreed_score = as_float(ws.cell(row=row_idx, column=15).value)

        is_total = title.startswith("ИТОГО ПО КОМПЕТЕНЦИИ")

        if is_total:
            competences.append({
                "name": title,
                "employee_score": employee_score,
                "manager_score": manager_score,
                "agreed_score": agreed_score,
                "indicators": indicators,
            })

            indicators = []
            continue

        indicators.append({
            "name": title,
            "employee_score": employee_score,
            "manager_score": manager_score,
            "agreed_score": agreed_score,
        })

    return competences


def parse_profile_sheet(wb, competences):
    if PROFILE_SHEET_NAME not in wb.sheetnames:
        return {
            "position": "",
            "real_profile": [],
            "target_profile": [],
        }

    ws = wb[PROFILE_SHEET_NAME]
    position = clean_text(ws["B2"].value)

    agreed_map = {item["name"]: item["agreed_score"] for item in competences}

    real_profile = []
    target_profile = []
    current_section = ""

    for row_idx in range(6, 15):
        section = clean_text(ws.cell(row=row_idx, column=1).value)
        name = clean_text(ws.cell(row=row_idx, column=2).value)

        if not name:
            continue

        if section:
            current_section = section

        real_value = agreed_map.get(name, as_float(ws.cell(row=row_idx, column=3).value))
        target_value = as_float(ws.cell(row=row_idx, column=4).value)

        real_profile.append({
            "name": name,
            "section": current_section,
            "value": real_value,
        })

        target_profile.append({
            "name": name,
            "section": current_section,
            "value": target_value,
        })

    return {
        "position": position,
        "real_profile": real_profile,
        "target_profile": target_profile,
    }


def parse_xlsx(file):
    raw = file.read()

    if not raw:
        raise ValueError("Файл пустой")

    wb = load_workbook(io.BytesIO(raw), data_only=True)

    competences = parse_agreed_sheet(wb)
    profile_data = parse_profile_sheet(wb, competences)

    return {
        "meta": {
            "position": profile_data["position"],
            "competences_count": len(competences),
        },
        "competences": [
            {"name": item["name"], "value": item["agreed_score"]}
            for item in competences
        ],
        "real_profile": profile_data["real_profile"],
        "target_profile": profile_data["target_profile"],
        "competency_details": competences,
    }